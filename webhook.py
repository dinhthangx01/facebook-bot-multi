from flask import Flask, request
import requests, os
from dotenv import load_dotenv
from openpyxl import load_workbook
import google.generativeai as genai
from waitress import serve

app = Flask(__name__)
load_dotenv()
VERIFY_TOKEN = "123abc"

PAGE_CONFIG, AI_COMMANDS = {}, {}

# =====================================================
# LOAD CONFIG FROM EXCEL (pages_config + ai_commands)
# =====================================================
def load_page_config():
    wb = load_workbook("pages_config.xlsx", read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        page_name, page_id, token, gemini_key, store_link, ai_prompt = row
        if page_id:
            PAGE_CONFIG[str(page_id)] = {
                "token": token,
                "gemini": gemini_key,
                "store": store_link or "",
                "prompt": ai_prompt or "You are a kind and helpful Heaven assistant. Reply naturally and briefly in English."
            }

def load_ai_commands():
    wb = load_workbook("ai_commands.xlsx", read_only=True)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        mode_key, trigger_keywords, desc, prompt = row
        if trigger_keywords:
            AI_COMMANDS[mode_key] = {
                "keywords": [k.strip().lower() for k in str(trigger_keywords).split(",")],
                "prompt": prompt or ""
            }

load_page_config()
load_ai_commands()

# =====================================================
# DETECT MODE BASED ON KEYWORDS
# =====================================================
def detect_mode(message):
    msg = message.lower()
    for mode, data in AI_COMMANDS.items():
        for keyword in data["keywords"]:
            if keyword in msg:
                return mode
    return "default"

# =====================================================
# SEND MESSAGE TO USER
# =====================================================
def send_message(user_id, text, page_id):
    token = PAGE_CONFIG[str(page_id)]["token"]
    url = f"https://graph.facebook.com/v17.0/me/messages?access_token={token}"
    payload = {"recipient": {"id": user_id}, "message": {"text": text}}
    requests.post(url, json=payload)

# =====================================================
# GENERATE GEMINI RESPONSE
# =====================================================
def generate_reply(message, api_key, base_prompt, mode_prompt=""):
    try:
        genai.configure(api_key=api_key)
        model = genai.GenerativeModel("gemini-2.0-flash")

        # Prompt kết hợp
        full_prompt = (
            f"{base_prompt}\n\n"
            f"{mode_prompt}\n\n"
            f"User message: '{message}'\n"
            f"Reply briefly in English, under 200 words."
        )

        response = model.generate_content(full_prompt)
        return response.text.strip() if response.text else "..."
    except Exception as e:
        print("⚠️ Gemini error:", e)
        return "Sorry, I’m having trouble replying right now."

# =====================================================
# VERIFY WEBHOOK
# =====================================================
@app.route("/webhook", methods=["GET"])
def verify():
    if request.args.get("hub.verify_token") == VERIFY_TOKEN:
        print("✅ Webhook verified successfully!")
        return request.args.get("hub.challenge"), 200
    return "Verification token mismatch", 403

# =====================================================
# HANDLE INCOMING MESSAGES
# =====================================================
@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.get_json()
    if data.get("object") == "page":
        for entry in data.get("entry", []):
            page_id = str(entry.get("id"))
            config = PAGE_CONFIG.get(page_id, {})
            gemini_key = config.get("gemini")
            store_link = config.get("store", "")
            base_prompt = config.get("prompt", "")

            for event in entry.get("messaging", []):
                if "message" in event and "text" in event["message"]:
                    sender_id = event["sender"]["id"]
                    user_msg = event["message"]["text"].strip()

                    mode = detect_mode(user_msg)
                    mode_prompt = AI_COMMANDS.get(mode, {}).get("prompt", "")

                    ai_reply = generate_reply(user_msg, gemini_key, base_prompt, mode_prompt)

                    # Nếu là chế độ bán hàng & có link cửa hàng
                    if mode == "buy_product" and store_link:
                        ai_reply += f"\n\n🛍️ Visit our Heaven Store:\n{store_link}"

                    send_message(sender_id, ai_reply, page_id)
    return "OK", 200

# =====================================================
# HOME
# =====================================================
@app.route("/", methods=["GET"])
def home():
    return "✅ HeavenBot multi-page (auto mode detect, English only, <200 words).", 200

# =====================================================
# START SERVER
# =====================================================
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    print(f"🚀 HeavenBot running on port {port}")
    serve(app, host="0.0.0.0", port=port)
